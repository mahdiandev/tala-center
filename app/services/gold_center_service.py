import copy
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from core.config import settings
from schemas.transaction import Transaction
from services.gold_price_service import GoldPriceService
from utils.shamsi import Shamsi
from utils.text import extract_digits


class GoldCenterService:
    """
    Manages the main Gold Center Excel database, processes deposit registrations,
    performs duplicate checks, and generates value-only export sheets using template file.
    """

    def __init__(self, gold_price_service: GoldPriceService) -> None:
        """
        Initializes the service with gold price lookup capabilities.

        Args:
            gold_price_service (GoldPriceService): The service to fetch historical gold prices.
        """
        self.gold_price_service = gold_price_service
        self.main_file_path = settings.gold_center_file

    def _is_numeric(self, val) -> bool:
        """
        Checks if a cell value represents a valid row number.

        Args:
            val: The cell value to inspect.

        Returns:
            bool: True if numeric, False otherwise.
        """
        if val is None:
            return False

        if isinstance(val, (int, float)):
            return True

        val_str = str(val).strip()
        return val_str.isdigit()

    def _copy_cell_style(self, source_cell, target_cell) -> None:
        """
        Copies font, fill, border, alignment, and number format from source to target.

        Args:
            source_cell: Openpyxl cell to copy formatting from.
            target_cell: Openpyxl cell to paste formatting to.
        """
        if source_cell.has_style:
            target_cell.font = copy.copy(source_cell.font)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.alignment = copy.copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format

    def _read_existing_transactions(self) -> set[tuple[int, str, str, str]]:
        """
        Reads the main Excel file and builds a set of unique existing transaction signatures.

        Each signature is a tuple of (amount, date_str, time_str, bank_name).

        Returns:
            set[tuple[int, str, str, str]]: Unique transaction signatures.

        Raises:
            FileNotFoundError: If the main Excel file does not exist.
        """
        if not self.main_file_path.exists():
            raise FileNotFoundError(f'Main gold center file not found at: {self.main_file_path}')

        signatures = set()
        wb = openpyxl.load_workbook(self.main_file_path, read_only=True, data_only=True)
        sheet: Worksheet = wb.active

        try:
            for row in sheet.iter_rows(min_row=2, max_col=31):
                if len(row) < 30:
                    continue

                amount_val = row[26].value
                if amount_val is None:
                    continue

                try:
                    amount_clean = int(extract_digits(str(amount_val)))
                    if amount_clean <= 0:
                        continue

                    date_str = str(row[27].value or '').strip()
                    time_str = str(row[28].value or '').strip()
                    bank_name = str(row[29].value or '').strip()

                    signatures.add((amount_clean, date_str, time_str, bank_name))
                except Exception:
                    continue
        finally:
            wb.close()

        return signatures

    def add_transactions(self, transactions: list[Transaction]) -> dict[str, int]:
        """
        Registers unique incoming transactions in the main Excel file.

        Resolves gold prices, applies required formulas, and appends rows.

        Args:
            transactions (list[Transaction]): List of incoming parsed transactions.

        Returns:
            dict[str, int]: Statistics of the registration process.

        Raises:
            FileNotFoundError: If the main Excel file does not exist.
            RuntimeError: If the main Excel file is locked or open in another application.
        """
        if not self.main_file_path.exists():
            raise FileNotFoundError(f'Main gold center file not found at: {self.main_file_path}')

        existing_signatures = self._read_existing_transactions()

        try:
            wb = openpyxl.load_workbook(self.main_file_path, read_only=False, data_only=False)
        except PermissionError as e:
            raise RuntimeError('فایل مرکز اصلی طلا باز هست. آن را ببندید و دوباره تلاش کنید..') from e

        sheet: Worksheet = wb.active

        added_count = 0
        skipped_count = 0
        bank_stats = {}

        try:
            for transaction in transactions:
                date_str = transaction.date.strftime('%Y/%m/%d')
                time_str = transaction.date.strftime('%H:%M:%S')

                signature = (transaction.amount, date_str, time_str, transaction.bank_name)

                if signature in existing_signatures:
                    skipped_count += 1
                    continue

                gold_price = self.gold_price_service.get_price(date_str)
                transaction.gold_price = gold_price

                next_row = sheet.max_row + 1

                if next_row == 2:
                    row_formula = '=1'
                else:
                    row_formula = f'=B{next_row-1}+1'

                def write_styled_cell(col_idx, value):
                    source_cell = sheet.cell(row=1, column=col_idx)
                    target_cell = sheet.cell(row=next_row, column=col_idx, value=value)
                    self._copy_cell_style(source_cell, target_cell)
                    return target_cell

                write_styled_cell(1, 2)
                write_styled_cell(2, row_formula)
                write_styled_cell(3, f'=AB{next_row}&" "&AC{next_row}')
                
                cell_d = write_styled_cell(4, 63752299629)
                cell_d.number_format = '0'

                write_styled_cell(9, 1)
                write_styled_cell(10, f'=AA{next_row}/M{next_row}')
                write_styled_cell(11, 1622)
                write_styled_cell(12, f'=ABS(J{next_row}*M{next_row}-AA{next_row})')
                write_styled_cell(13, transaction.gold_price)
                
                cell_n = write_styled_cell(14, 2720000260362)
                cell_n.number_format = '0'

                write_styled_cell(16, 0)
                write_styled_cell(17, 0)
                write_styled_cell(18, 0)
                write_styled_cell(19, 0)
                write_styled_cell(27, transaction.amount)
                write_styled_cell(28, date_str)
                write_styled_cell(29, time_str)
                write_styled_cell(30, transaction.bank_name)
                write_styled_cell(31, None)

                added_count += 1
                bank_stats[transaction.bank_name] = bank_stats.get(transaction.bank_name, 0) + 1

            if added_count > 0:
                try:
                    wb.save(self.main_file_path)
                except PermissionError as e:
                    raise RuntimeError('فایل مرکز اصلی طلا باز هست. آن را ببندید و دوباره تلاش کنید..') from e
        finally:
            wb.close()

        return {
            'added': added_count,
            'skipped': skipped_count,
            'bank_stats': bank_stats
        }

    def export_transactions(self) -> int:
        """
        Generates a styled value-only Excel snapshot of non-exported transactions using a pre-defined template.

        Employs a dual-loading mechanism:
        - Reads evaluated formula outputs under a data_only context with fallback Python calculations.
        - Writes back the export status under a formula-preserving write context.
        Excludes the trailing 5 metadata columns (AA to AE) from the export.

        Returns:
            int: The total count of exported transactions.

        Raises:
            FileNotFoundError: If the main Excel file or template file does not exist.
            RuntimeError: If the main Excel file is locked or open in another application.
        """
        if not self.main_file_path.exists():
            raise FileNotFoundError(f'Main gold center file not found at: {self.main_file_path}')

        template_path = Path(__file__).resolve().parent.parent / 'templates' / 'قالب خروجی.xlsx'
        if not template_path.exists():
            raise FileNotFoundError(f'Export template file not found at: {template_path}')

        wb_export = openpyxl.load_workbook(template_path, read_only=False, data_only=False)
        sheet_export: Worksheet = wb_export.active

        wb_main_data = openpyxl.load_workbook(self.main_file_path, read_only=True, data_only=True)
        sheet_main_data: Worksheet = wb_main_data.active

        try:
            wb_main_write = openpyxl.load_workbook(self.main_file_path, read_only=False, data_only=False)
        except PermissionError as e:
            wb_main_data.close()
            wb_export.close()
            raise RuntimeError('فایل مرکز اصلی طلا باز هست. آن را ببندید و دوباره تلاش کنید..') from e

        sheet_main_write: Worksheet = wb_main_write.active

        exported_count = 0

        try:
            for r in range(2, sheet_main_data.max_row + 1):
                row_idx_data = sheet_main_data.cell(row=r, column=2).value
                row_idx_write = sheet_main_write.cell(row=r, column=2).value
                amount_val = sheet_main_data.cell(row=r, column=27).value
                export_status = sheet_main_data.cell(row=r, column=31).value

                is_row_idx_valid = False
                if row_idx_data is not None and self._is_numeric(row_idx_data):
                    is_row_idx_valid = True
                elif row_idx_write is not None and str(row_idx_write).strip().startswith('='):
                    is_row_idx_valid = True

                if not is_row_idx_valid or amount_val is None:
                    continue

                if export_status == 'خروجی گرفته شد':
                    continue

                exported_count += 1
                next_export_row = exported_count + 1

                amount_int = int(extract_digits(str(amount_val)))
                date_str = str(sheet_main_data.cell(row=r, column=28).value or '').strip()
                time_str = str(sheet_main_data.cell(row=r, column=29).value or '').strip()
                gold_price = int(sheet_main_data.cell(row=r, column=13).value or 0)

                datetime_str = f'{date_str} {time_str}'.strip()
                gold_weight = amount_int / gold_price if gold_price > 0 else 0
                discount = int(abs(gold_weight * gold_price - amount_int))

                for col_idx in range(1, 27):
                    val = sheet_main_data.cell(row=r, column=col_idx).value

                    if val is None:
                        if col_idx == 1:
                            val = 2
                        elif col_idx == 2:
                            val = r - 1
                        elif col_idx == 3:
                            val = datetime_str
                        elif col_idx == 4:
                            val = 63752299629
                        elif col_idx == 9:
                            val = 1
                        elif col_idx == 10:
                            val = gold_weight
                        elif col_idx == 11:
                            val = 1622
                        elif col_idx == 12:
                            val = discount
                        elif col_idx == 13:
                            val = gold_price
                        elif col_idx == 14:
                            val = 2720000260362
                        elif col_idx in (16, 17, 18, 19):
                            val = 0

                    target_cell = sheet_export.cell(row=next_export_row, column=col_idx, value=val)

                    if col_idx in (4, 14):
                        target_cell.number_format = '0'

                sheet_main_write.cell(row=r, column=31, value='خروجی گرفته شد')

            if exported_count > 0:
                time_stamp = Shamsi.now().strftime('%Y-%m-%d-%H-%M-%S')
                export_filename = f'tala-export-{time_stamp}.xlsx'
                export_file_path = settings.output_folder / export_filename

                settings.output_folder.mkdir(parents=True, exist_ok=True)
                wb_export.save(export_file_path)
                try:
                    wb_main_write.save(self.main_file_path)
                except PermissionError as e:
                    raise RuntimeError('فایل مرکز اصلی طلا باز هست. آن را ببندید و دوباره تلاش کنید..') from e

        finally:
            wb_main_data.close()
            wb_main_write.close()
            wb_export.close()

        return exported_count
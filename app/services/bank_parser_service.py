import tempfile
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill
import xlrd
from schemas.transaction import Transaction
from utils.shamsi import Shamsi
from utils.text import normalize_text, extract_digits


class BankParserService:
    """
    Identifies the bank type of financial statement files and extracts deposit transactions.
    Supports both legacy .xls and modern .xlsx formats.
    """

    def __init__(self) -> None:
        """
        Initializes the parser service.
        """
        pass

    def _is_numeric(self, val) -> bool:
        """
        Checks if a cell value represents a valid row number.
        """
        if val is None:
            return False

        if isinstance(val, (int, float)):
            return True

        val_str = str(val).strip()
        return val_str.isdigit()

    def _format_time_value(self, val) -> str:
        """
        Converts any cell value representing time into a valid HH:MM:SS string.

        Supports datetime.time/datetime.datetime objects, float strings, and floats.

        Args:
            val: The raw cell value.

        Returns:
            str: The formatted time string.
        """
        if val is None:
            return ''

        if hasattr(val, 'strftime'):
            return val.strftime('%H:%M:%S')

        val_str = str(val).strip()

        try:
            float_val = float(val_str)
            if 0.0 <= float_val <= 1.0:
                total_seconds = int(round(float_val * 86400.0))
                hours = total_seconds // 3600
                minutes = (total_seconds % 3600) // 60
                seconds = total_seconds % 60
                hours = hours % 24
                return f'{hours:02d}:{minutes:02d}:{seconds:02d}'
        except ValueError:
            pass

        return val_str

    def _is_cumulative(self, desc_val) -> bool:
        """
        Detects if a description contains cumulative keyword indicators.
        """
        if desc_val is None:
            return False

        normalized = normalize_text(str(desc_val))
        return 'شاپرک' in normalized or 'تجمیعی' in normalized

    def _convert_xls_to_xlsx(self, xls_path: Path, xlsx_path: Path) -> None:
        """
        Reads a legacy .xls file, sanitizes data types, and converts it into .xlsx.
        Tries to use Excel itself via pywin32 to preserve styles/images, 
        and falls back to xlrd + openpyxl if unavailable.

        Args:
            xls_path (Path): Path to the legacy .xls file.
            xlsx_path (Path): Destination path for the converted .xlsx file.
        """
        excel = None
        wb = None
        try:
            import win32com.client as win32
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.DisplayAlerts = False
            wb = excel.Workbooks.Open(str(xls_path.absolute()))
            wb.SaveAs(str(xlsx_path.absolute()), FileFormat=51)
            wb.Close(False)
            excel.Application.Quit()
            return
        except Exception:
            if wb:
                try:
                    wb.Close(False)
                except Exception:
                    pass
            if excel:
                try:
                    excel.Application.Quit()
                except Exception:
                    pass

        xls_book = xlrd.open_workbook(str(xls_path))
        xlsx_book = openpyxl.Workbook()

        first_sheet = True
        for sheet_name in xls_book.sheet_names():
            xls_sheet = xls_book.sheet_by_name(sheet_name)

            if first_sheet:
                xlsx_sheet = xlsx_book.active
                xlsx_sheet.title = sheet_name
                first_sheet = False
            else:
                xlsx_sheet = xlsx_book.create_sheet(title=sheet_name)

            for r in range(xls_sheet.nrows):
                for c in range(xls_sheet.ncols):
                    cell_value = xls_sheet.cell_value(r, c)

                    if isinstance(cell_value, float) and cell_value.is_integer():
                        cell_value = int(cell_value)

                    xlsx_sheet.cell(row=r + 1, column=c + 1, value=cell_value)

        xlsx_book.save(xlsx_path)
        xlsx_book.close()

    def _apply_modifications_via_excel(self, file_path: Path, target_path: Path, modifications: list[dict], is_xls: bool) -> Path:
        """
        Applies row deletions, insertions, and updates natively using Microsoft Excel
        via win32com to guarantee 100% preservation of styles, fonts, and images.

        Args:
            file_path (Path): Path to the original Excel file.
            target_path (Path): Path to the converted temporary Excel file.
            modifications (list[dict]): Collected splitting modifications.
            is_xls (bool): True if the original file format is legacy .xls.

        Returns:
            Path: The modified and saved Excel file path.
        """
        excel = None
        wb = None
        try:
            import win32com.client as win32
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            excel.DisplayAlerts = False
            excel.Visible = False

            wb = excel.Workbooks.Open(str(target_path.absolute()))
            sheet = wb.ActiveSheet
            max_cols = sheet.UsedRange.Columns.Count

            modifications.sort(key=lambda x: x['row_index'], reverse=True)

            for mod in modifications:
                r = mod['row_index']
                split_amounts = mod['split_amounts']
                amount_col_idx = mod['amount_col_idx']
                desc_col_idx = mod['desc_col_idx']
                time_col_idx = mod['time_col_idx']
                time_str = mod['time_str']

                sheet.Rows(r).Copy()

                for _ in range(len(split_amounts)):
                    sheet.Rows(r + 1).Insert(Shift=-4121)

                original_height = sheet.Rows(r).RowHeight

                for i in range(len(split_amounts)):
                    current_row = r + 1 + i
                    sheet.Rows(r).Copy()
                    sheet.Rows(current_row).PasteSpecial(Paste=-4104)

                    sheet.Rows(current_row).RowHeight = original_height

                    sheet.Cells(current_row, amount_col_idx).Value = split_amounts[i]
                    sheet.Cells(current_row, amount_col_idx).NumberFormat = '#,##0'
                    sheet.Cells(current_row, desc_col_idx).Value = 'ثبت شده توسط کاربر'
                    sheet.Rows(current_row).Interior.Color = 13434879

                    try:
                        h, m, s = map(int, time_str.split(':'))
                        s = s + i
                        m = m + s // 60
                        s = s % 60
                        h = (h + m // 60) % 24
                        m = m % 60
                        new_time_str = f'{h:02d}:{m:02d}:{s:02d}'
                    except Exception:
                        new_time_str = time_str

                    sheet.Cells(current_row, time_col_idx).Value = new_time_str

                sheet.Rows(r).Delete()

            excel.CutCopyMode = False

            if is_xls:
                final_xlsx_path = file_path.with_suffix('.xlsx')
                wb.SaveAs(str(final_xlsx_path.absolute()), FileFormat=51)
                wb.Close(False)
                excel.Application.Quit()
                if target_path.exists():
                    try:
                        target_path.unlink()
                    except OSError:
                        pass
                if file_path.exists():
                    try:
                        file_path.unlink()
                    except OSError:
                        pass
                return final_xlsx_path
            else:
                wb.Save()
                wb.Close(False)
                excel.Application.Quit()
                return file_path

        except Exception as e:
            if wb:
                try:
                    wb.Close(False)
                except Exception:
                    pass
            if excel:
                try:
                    excel.Application.Quit()
                except Exception:
                    pass
            print(f'[BankParser Excel Error] Failed to modify via win32com: {str(e)}. Switched to openpyxl fallback.')

            wb = openpyxl.load_workbook(target_path, read_only=False, data_only=False)
            sheet = wb.active
            yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            modifications.sort(key=lambda x: x['row_index'], reverse=True)

            for mod in modifications:
                r = mod['row_index']
                split_amounts = mod['split_amounts']
                amount_col_idx = mod['amount_col_idx']
                desc_col_idx = mod['desc_col_idx']
                time_col_idx = mod['time_col_idx']
                time_str = mod['time_str']

                cumulative_values = [sheet.cell(row=r, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
                original_height = sheet.row_dimensions[r].height

                sheet.delete_rows(r, 1)
                sheet.insert_rows(r, len(split_amounts))

                for i, split_amount in enumerate(split_amounts):
                    current_row = r + i
                    sheet.row_dimensions[current_row].height = original_height

                    for col_idx, val in enumerate(cumulative_values, start=1):
                        sheet.cell(row=current_row, column=col_idx, value=val)

                    sheet.cell(row=current_row, column=amount_col_idx, value=split_amount).number_format = '#,##0'
                    sheet.cell(row=current_row, column=desc_col_idx, value='ثبت شده توسط کاربر')

                    try:
                        h, m, s = map(int, time_str.split(':'))
                        s = s + i
                        m = m + s // 60
                        s = s % 60
                        h = (h + m // 60) % 24
                        m = m % 60
                        new_time_str = f'{h:02d}:{m:02d}:{s:02d}'
                    except Exception:
                        new_time_str = time_str

                    sheet.cell(row=current_row, column=time_col_idx, value=new_time_str)

                    for col_idx in range(1, len(cumulative_values) + 1):
                        sheet.cell(row=current_row, column=col_idx).fill = yellow_fill

            wb.save(target_path)
            wb.close()

            if is_xls:
                final_xlsx_path = file_path.with_suffix('.xlsx')
                import shutil
                try:
                    if target_path.exists():
                        shutil.copy2(target_path, final_xlsx_path)
                        target_path.unlink()
                    if file_path.exists():
                        file_path.unlink()
                except OSError:
                    pass
                return final_xlsx_path
            else:
                return file_path

    def _detect_bank_name(self, sheet: Worksheet) -> str | None:
        """
        Detects the bank name directly from the active Worksheet memory buffer.

        Args:
            sheet (Worksheet): The active worksheet in RAM.

        Returns:
            str | None: The bank name, or None if unrecognized.
        """
        h4_val = sheet.cell(row=4, column=8).value
        if h4_val and '119727908005' in str(h4_val):
            return 'ملی'

        e2_val = sheet.cell(row=2, column=5).value
        if e2_val and '2719-830-2221768-1' in str(e2_val):
            return 'ایران زمین'

        g15_val = sheet.cell(row=15, column=7).value
        if g15_val and '0279026634107' in str(g15_val):
            return 'تجارت'

        l2_val = sheet.cell(row=2, column=12).value
        if l2_val and '120703703000' in str(l2_val):
            return 'صادرات'

        return None

    def parse_excel(self, file_path: Path, on_cumulative_found=None) -> list[Transaction]:
        """
        Parses the bank statement Excel file and returns normalized deposit transactions.

        Exactly resolves detection and parsing within a single file open-close cycle.

        Args:
            file_path (Path): Path to the bank statement Excel file.
            on_cumulative_found: Callback triggered when a cumulative transaction is met.

        Returns:
            list[Transaction]: Extracted deposits.
        """
        if file_path.name.startswith('~$'):
            print(f'[BankParser] Skipping Excel temporary lock file: {file_path.name}')
            return []

        print(f'[BankParser] Starting parsing for file: {file_path.name}')
        is_xls = file_path.suffix.lower() == '.xls'
        target_path = file_path

        if is_xls:
            print(f'[BankParser] Conversion triggered for legacy XLS: {file_path.name}')
            target_path = Path(tempfile.gettempdir()) / file_path.with_suffix('.temp.xlsx').name
            self._convert_xls_to_xlsx(file_path, target_path)

        transactions = []
        modifications = []

        try:
            wb = openpyxl.load_workbook(target_path, read_only=False, data_only=True)
            sheet: Worksheet = wb.active

            try:
                bank_name = self._detect_bank_name(sheet)
                print(f'[BankParser] Detection completed. Bank name: {bank_name}')

                if not bank_name:
                    raise ValueError(f'Bank statement type unrecognized for file: {file_path.name}')

                if bank_name == 'ملی':
                    transactions, modifications = self._parse_melli(sheet, on_cumulative_found)
                elif bank_name == 'ایران زمین':
                    transactions, modifications = self._parse_iranzamin(sheet, on_cumulative_found)
                elif bank_name == 'تجارت':
                    transactions, modifications = self._parse_tejarat(sheet, on_cumulative_found)
                elif bank_name == 'صادرات':
                    transactions, modifications = self._parse_saderat(sheet, on_cumulative_found)
            finally:
                if not modifications:
                    wb.close()
        finally:
            if not modifications:
                if is_xls and target_path.exists():
                    try:
                        target_path.unlink()
                    except OSError:
                        pass

        if modifications:
            modified_path = self._apply_modifications_via_excel(file_path, target_path, modifications, is_xls)
            if is_xls:
                final_xlsx_path = file_path.with_suffix('.xlsx')
                return self.parse_excel(final_xlsx_path, on_cumulative_found)
            else:
                return self.parse_excel(file_path, on_cumulative_found)

        return transactions

    def _parse_melli(self, sheet: Worksheet, on_cumulative_found=None) -> tuple[list[Transaction], list[dict]]:
        """
        Extracts deposit rows from Melli Bank sheet.
        """
        transactions = []
        modifications = []
        row_count = 0
        for row in sheet.iter_rows(min_row=1, max_col=9):
            row_count += 1
            if len(row) < 9:
                continue

            row_idx = row[0].value
            transaction_type = row[4].value
            if not self._is_numeric(row_idx) or not transaction_type:
                continue

            normalized_type = normalize_text(str(transaction_type))
            if normalized_type != 'واریز':
                continue

            try:
                amount_raw = row[5].value
                if amount_raw is None:
                    continue

                if isinstance(amount_raw, (int, float)):
                    amount_val = int(amount_raw)
                else:
                    clean_str = str(amount_raw).split('.')[0]
                    amount_str = extract_digits(clean_str)
                    amount_val = int(amount_str) if amount_str else 0

                if amount_val <= 0:
                    continue

                desc_val = row[8].value
                if self._is_cumulative(desc_val):
                    date_str = str(row[1].value or '').strip()
                    time_str = self._format_time_value(row[2].value)
                    if on_cumulative_found:
                        split_amounts = on_cumulative_found('ملی', amount_val, date_str, time_str)
                        if split_amounts:
                            modifications.append({
                                'row_index': row_count,
                                'split_amounts': split_amounts,
                                'amount_col_idx': 6,
                                'desc_col_idx': 9,
                                'time_col_idx': 3,
                                'time_str': time_str
                            })
                    continue

                date_str = str(row[1].value or '').strip()
                time_str = self._format_time_value(row[2].value)
                dt_str = f'{date_str} {time_str}'.strip()
                dt_obj = Shamsi.from_str(dt_str)

                transactions.append(
                    Transaction(
                        amount=amount_val,
                        date=dt_obj,
                        bank_name='ملی'
                    )
                )
            except Exception as e:
                if 'تراکنش تجمیعی' in str(e) or 'متوقف شد' in str(e):
                    raise e
                print(f'[Melli Debug] Row {row_count} failed with exception: {str(e)}')
                continue

        return transactions, modifications

    def _parse_iranzamin(self, sheet: Worksheet, on_cumulative_found=None) -> tuple[list[Transaction], list[dict]]:
        """
        Extracts deposit rows from Iran Zamin Bank sheet.
        """
        transactions = []
        modifications = []
        row_count = 0
        for row in sheet.iter_rows(min_row=1, max_col=5):
            row_count += 1
            if len(row) < 5:
                continue

            row_idx = row[0].value
            if not self._is_numeric(row_idx):
                continue

            try:
                amount_raw = row[4].value
                if amount_raw is None:
                    continue

                if isinstance(amount_raw, (int, float)):
                    amount_val = int(amount_raw)
                else:
                    clean_str = str(amount_raw).split('.')[0]
                    amount_str = extract_digits(clean_str)
                    amount_val = int(amount_str) if amount_str else 0

                if amount_val <= 0:
                    continue

                desc_val = row[1].value
                if self._is_cumulative(desc_val):
                    date_str = str(row[2].value or '').strip()
                    time_str = self._format_time_value(row[3].value)
                    if on_cumulative_found:
                        split_amounts = on_cumulative_found('ایران زمین', amount_val, date_str, time_str)
                        if split_amounts:
                            modifications.append({
                                'row_index': row_count,
                                'split_amounts': split_amounts,
                                'amount_col_idx': 5,
                                'desc_col_idx': 2,
                                'time_col_idx': 4,
                                'time_str': time_str
                            })
                    continue

                date_str = str(row[2].value or '').strip()
                time_str = self._format_time_value(row[3].value)
                dt_str = f'{date_str} {time_str}'.strip()
                dt_obj = Shamsi.from_str(dt_str)

                transactions.append(
                    Transaction(
                        amount=amount_val,
                        date=dt_obj,
                        bank_name='ایران زمین'
                    )
                )
            except Exception as e:
                if 'تراکنش تجمیعی' in str(e) or 'متوقف شد' in str(e):
                    raise e
                print(f'[Iran Zamin Debug] Row {row_count} failed with exception: {str(e)}')
                continue

        return transactions, modifications

    def _parse_tejarat(self, sheet: Worksheet, on_cumulative_found=None) -> tuple[list[Transaction], list[dict]]:
        """
        Extracts deposit rows from Tejarat Bank sheet.
        """
        transactions = []
        modifications = []
        row_count = 0
        for row in sheet.iter_rows(min_row=1, max_col=16):
            row_count += 1
            if len(row) < 16:
                continue

            row_idx = row[15].value
            if not self._is_numeric(row_idx):
                continue

            try:
                amount_raw = row[10].value
                if amount_raw is None:
                    continue

                if isinstance(amount_raw, (int, float)):
                    amount_val = int(amount_raw)
                else:
                    clean_str = str(amount_raw).split('.')[0]
                    amount_str = extract_digits(clean_str)
                    amount_val = int(amount_str) if amount_str else 0

                if amount_val <= 0:
                    continue

                desc_val = row[7].value
                if self._is_cumulative(desc_val):
                    date_str = str(row[14].value or '').strip()
                    time_str = self._format_time_value(row[13].value)
                    if on_cumulative_found:
                        split_amounts = on_cumulative_found('تجارت', amount_val, date_str, time_str)
                        if split_amounts:
                            modifications.append({
                                'row_index': row_count,
                                'split_amounts': split_amounts,
                                'amount_col_idx': 11,
                                'desc_col_idx': 8,
                                'time_col_idx': 14,
                                'time_str': time_str
                            })
                    continue

                date_str = str(row[14].value or '').strip()
                time_str = self._format_time_value(row[13].value)
                dt_str = f'{date_str} {time_str}'.strip()
                dt_obj = Shamsi.from_str(dt_str)

                transactions.append(
                    Transaction(
                        amount=amount_val,
                        date=dt_obj,
                        bank_name='تجارت'
                    )
                )
            except Exception as e:
                if 'تراکنش تجمیعی' in str(e) or 'متوقف شد' in str(e):
                    raise e
                print(f'[Tejarat Debug] Row {row_count} failed with exception: {str(e)}')
                continue

        return transactions, modifications

    def _parse_saderat(self, sheet: Worksheet, on_cumulative_found=None) -> tuple[list[Transaction], list[dict]]:
        """
        Extracts deposit rows from Saderat Bank sheet.
        """
        transactions = []
        modifications = []
        row_count = 0
        for row in sheet.iter_rows(min_row=1, max_col=9):
            row_count += 1
            if len(row) < 9:
                continue

            row_idx = row[0].value
            if not self._is_numeric(row_idx):
                continue

            try:
                amount_raw = row[8].value
                if amount_raw is None:
                    continue

                if isinstance(amount_raw, (int, float)):
                    amount_val = int(amount_raw)
                else:
                    clean_str = str(amount_raw).split('.')[0]
                    amount_str = extract_digits(clean_str)
                    amount_val = int(amount_str) if amount_str else 0

                if amount_val <= 0:
                    continue

                desc_val = row[3].value
                if self._is_cumulative(desc_val):
                    date_str = str(row[1].value or '').strip()
                    time_str = self._format_time_value(row[2].value)
                    if on_cumulative_found:
                        split_amounts = on_cumulative_found('صادرات', amount_val, date_str, time_str)
                        if split_amounts:
                            modifications.append({
                                'row_index': row_count,
                                'split_amounts': split_amounts,
                                'amount_col_idx': 9,
                                'desc_col_idx': 4,
                                'time_col_idx': 3,
                                'time_str': time_str
                            })
                    continue

                date_str = str(row[1].value or '').strip()
                time_str = self._format_time_value(row[2].value)
                dt_str = f'{date_str} {time_str}'.strip()
                dt_obj = Shamsi.from_str(dt_str)

                transactions.append(
                    Transaction(
                        amount=amount_val,
                        date=dt_obj,
                        bank_name='صادرات'
                    )
                )
            except Exception as e:
                if 'تراکنش تجمیعی' in str(e) or 'متوقف شد' in str(e):
                    raise e
                print(f'[Saderat Debug] Row {row_count} failed with exception: {str(e)}')
                continue

        return transactions, modifications
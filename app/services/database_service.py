import copy
import shutil
from pathlib import Path
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from core.config import settings
from schemas.transaction import Transaction
from services.gold_price_service import GoldPriceService
from utils.shamsi import Shamsi
from utils.text import normalize_text, extract_digits


class DatabaseService:
    """
    Manages the main database, processes deposit registrations,
    performs duplicate checks, and generates value-only export sheets.
    """

    def __init__(self, gold_price_service: GoldPriceService) -> None:
        """
        Initializes the service with gold price lookup capabilities.

        Args:
            gold_price_service (GoldPriceService): The service to fetch gold prices.
        """
        self.gold_price_service = gold_price_service
        self.main_file_path = settings.database_file

    def _is_numeric(self, val) -> bool:
        """
        Checks if a cell value represents a valid row number.

        Args:
            val: The cell value to inspect.

        Returns:
            bool: True if numeric, False otherwise.
        """
        if val is None:
            return False

        if isinstance(val, (int, float)):
            return True

        val_str = str(val).strip()
        return val_str.isdigit()

    def _normalize_date_str(self, date_val) -> str:
        """
        Normalizes a date value into YYYY/MM/DD format.

        Args:
            date_val: The raw date cell value.

        Returns:
            str: The normalized date string.
        """
        if date_val is None:
            return ''

        if hasattr(date_val, 'strftime'):
            return date_val.strftime('%Y/%m/%d')

        val_str = str(date_val).strip().replace('-', '/')
        parts = val_str.split('/')

        if len(parts) == 3:
            try:
                y = int(parts[0])
                m = int(parts[1])
                d = int(parts[2])
                return f'{y:04d}/{m:02d}/{d:02d}'
            except ValueError:
                pass

        return val_str

    def _normalize_time_str(self, time_val) -> str:
        """
        Normalizes a time value into HH:MM:SS format.

        Args:
            time_val: The raw time cell value.

        Returns:
            str: The normalized time string.
        """
        if time_val is None:
            return ''

        if hasattr(time_val, 'strftime'):
            return time_val.strftime('%H:%M:%S')

        val_str = str(time_val).strip()
        parts = val_str.split(':')

        if len(parts) == 3:
            try:
                h = int(parts[0])
                m = int(parts[1])
                s = int(parts[2])
                return f'{h:02d}:{m:02d}:{s:02d}'
            except ValueError:
                pass

        return val_str

    def _copy_cell_style(self, source_cell, target_cell) -> None:
        """
        Copies style properties from a source cell to a target cell.

        Args:
            source_cell: Openpyxl cell to copy formatting from.
            target_cell: Openpyxl cell to paste formatting to.
        """
        if source_cell.has_style:
            target_cell.font = copy.copy(source_cell.font)
            target_cell.fill = copy.copy(source_cell.fill)
            target_cell.border = copy.copy(source_cell.border)
            target_cell.alignment = copy.copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format

    def _read_existing_transactions(self) -> set[tuple[int, str, str, str]]:
        """
        Reads the main Excel file and builds a set of unique existing transaction signatures.

        Each signature is a tuple of (amount, date_str, time_str, bank_name).

        Returns:
            set[tuple[int, str, str, str]]: Unique transaction signatures.

        Raises:
            FileNotFoundError: If the main Excel file does not exist.
        """
        if not self.main_file_path.exists():
            raise FileNotFoundError(f'Main database file not found at: {self.main_file_path}')

        signatures = set()
        wb = openpyxl.load_workbook(self.main_file_path, read_only=True, data_only=True)
        sheet: Worksheet = wb.active

        try:
            for row in sheet.iter_rows(min_row=2, max_col=30):
                if len(row) < 30:
                    continue

                amount_val = row[26].value
                if amount_val is None:
                    continue

                try:
                    amount_clean = int(extract_digits(str(amount_val)))
                    if amount_clean <= 0:
                        continue

                    raw_date = row[27].value
                    raw_time = row[28].value
                    raw_bank = row[29].value

                    date_clean = self._normalize_date_str(raw_date)
                    time_clean = self._normalize_time_str(raw_time)
                    bank_clean = normalize_text(str(raw_bank or '').strip())

                    signatures.add((amount_clean, date_clean, time_clean, bank_clean))
                except Exception:
                    continue
        finally:
            wb.close()

        return signatures

    def _create_database_backup(self) -> None:
        """
        Creates a timestamped backup of the main database in settings.backup_folder
        and maintains a maximum of 10 backup files.
        """
        backup_dir = settings.backup_folder
        if not backup_dir or not self.main_file_path.exists():
            return

        timestamp = Shamsi.now().strftime('%Y-%m-%d-%H-%M-%S')
        backup_path = backup_dir / f'database-backup-{timestamp}.xlsx'

        try:
            shutil.copy2(self.main_file_path, backup_path)
        except OSError:
            return

        try:
            backup_files = list(backup_dir.glob('database-backup-*.xlsx'))
            if len(backup_files) > 10:
                backup_files.sort(key=lambda x: x.stat().st_mtime)
                for old_file in backup_files[:-10]:
                    try:
                        old_file.unlink()
                    except OSError:
                        pass
        except Exception:
            pass

    def add_transactions(self, transactions: list[Transaction]) -> dict[str, int]:
        """
        Registers unique incoming transactions in the main Excel database file.

        Resolves gold prices, applies required formulas, and appends rows.

        Args:
            transactions (list[Transaction]): List of incoming parsed transactions.

        Returns:
            dict[str, int]: Statistics of the registration process.

        Raises:
            FileNotFoundError: If the main Excel database file does not exist.
            RuntimeError: If the main Excel file is locked or open in another application.
        """
        if not self.main_file_path.exists():
            raise FileNotFoundError(f'Main database file not found at: {self.main_file_path}')

        existing_signatures = self._read_existing_transactions()

        try:
            wb = openpyxl.load_workbook(self.main_file_path, read_only=False, data_only=False)
        except PermissionError as e:
            raise RuntimeError('فایل مرکز اصلی طلا باز هست. آن را ببندید و دوباره تلاش کنید..') from e

        sheet: Worksheet = wb.active

        added_count = 0
        skipped_count = 0
        bank_stats = {}

        try:
            for transaction in transactions:
                date_str = transaction.date.strftime('%Y/%m/%d')
                time_str = transaction.date.strftime('%H:%M:%S')

                normalized_date = self._normalize_date_str(date_str)
                normalized_time = self._normalize_time_str(time_str)
                normalized_bank = normalize_text(transaction.bank_name)

                signature = (transaction.amount, normalized_date, normalized_time, normalized_bank)

                if signature in existing_signatures:
                    skipped_count += 1
                    continue

                gold_price = self.gold_price_service.get_price(date_str)
                transaction.gold_price = gold_price

                next_row = sheet.max_row + 1

                if next_row == 2:
                    row_formula = '=1'
                else:
                    row_formula = f'=B{next_row-1}+1'

                def write_styled_cell(col_idx, value):
                    source_cell = sheet.cell(row=1, column=col_idx)
                    target_cell = sheet.cell(row=next_row, column=col_idx, value=value)
                    self._copy_cell_style(source_cell, target_cell)
                    return target_cell

                write_styled_cell(1, 2)
                write_styled_cell(2, row_formula)
                write_styled_cell(3, f'=AB{next_row}&" "&AC{next_row}')
                
                cell_d = write_styled_cell(4, 63752299629)
                cell_d.number_format = '0'

                write_styled_cell(9, 1)
                write_styled_cell(10, f'=AA{next_row}/M{next_row}')
                write_styled_cell(11, 1622)
                write_styled_cell(12, f'=ABS(J{next_row}*M{next_row}-AA{next_row})')
                write_styled_cell(13, transaction.gold_price)
                
                cell_n = write_styled_cell(14, 2720000260362)
                cell_n.number_format = '0'

                write_styled_cell(16, 0)
                write_styled_cell(17, 0)
                write_styled_cell(18, 0)
                write_styled_cell(19, 0)
                write_styled_cell(27, transaction.amount)
                write_styled_cell(28, date_str)
                write_styled_cell(29, time_str)
                write_styled_cell(30, transaction.bank_name)
                write_styled_cell(31, None)

                added_count += 1
                bank_stats[transaction.bank_name] = bank_stats.get(transaction.bank_name, 0) + 1

            if added_count > 0:
                try:
                    wb.save(self.main_file_path)
                    self._create_database_backup()
                except PermissionError as e:
                    raise RuntimeError('فایل مرکز اصلی طلا باز هست. آن را ببندید و دوباره تلاش کنید..') from e
        finally:
            wb.close()

        return {
            'added': added_count,
            'skipped': skipped_count,
            'bank_stats': bank_stats
        }

    def export_transactions(self) -> int:
        """
        Generates a styled value-only Excel snapshot of non-exported transactions.

        Employs a template file and writes back export statuses.

        Returns:
            int: The total count of exported transactions.

        Raises:
            FileNotFoundError: If the main Excel database file or template file does not exist.
            RuntimeError: If the main Excel file is locked or open in another application.
        """
        if not self.main_file_path.exists():
            raise FileNotFoundError(f'Main database file not found at: {self.main_file_path}')

        template_path = Path(__file__).resolve().parent.parent / 'templates' / 'export.xlsx'
        if not template_path.exists():
            raise FileNotFoundError(f'Export template file not found at: {template_path}')

        try:
            wb_main = openpyxl.load_workbook(self.main_file_path, read_only=False, data_only=False)
        except PermissionError as e:
            raise RuntimeError('فایل مرکز اصلی طلا باز هست. آن را ببندید و دوباره تلاش کنید..') from e

        sheet_main: Worksheet = wb_main.active
        eligible_rows = []

        for r in range(2, sheet_main.max_row + 1):
            row_idx_val = sheet_main.cell(row=r, column=2).value
            amount_val = sheet_main.cell(row=r, column=27).value
            export_status = sheet_main.cell(row=r, column=31).value

            is_row_valid = False
            if row_idx_val is not None:
                if self._is_numeric(row_idx_val):
                    is_row_valid = True
                elif isinstance(row_idx_val, str) and row_idx_val.strip().startswith('='):
                    is_row_valid = True

            if not is_row_valid or amount_val is None:
                continue

            if export_status == 'خروجی گرفته شد':
                continue

            eligible_rows.append(r)

        if not eligible_rows:
            wb_main.close()
            return 0

        start_cell_val = sheet_main.cell(row=2, column=2).value
        start_num = 1
        if start_cell_val is not None:
            try:
                digits_str = extract_digits(str(start_cell_val))
                if digits_str:
                    start_num = int(digits_str)
            except Exception:
                start_num = 1

        chunk_size = 299
        chunks = [eligible_rows[i:i + chunk_size] for i in range(0, len(eligible_rows), chunk_size)]
        time_stamp = Shamsi.now().strftime('%Y-%m-%d-%H-%M-%S')
        settings.output_folder.mkdir(parents=True, exist_ok=True)
        exported_count = 0

        try:
            for chunk_idx, chunk in enumerate(chunks, start=1):
                wb_export = openpyxl.load_workbook(template_path, read_only=False, data_only=False)
                sheet_export: Worksheet = wb_export.active

                for local_idx, r in enumerate(chunk):
                    amount_val = sheet_main.cell(row=r, column=27).value
                    amount_int = int(extract_digits(str(amount_val)))
                    date_str = str(sheet_main.cell(row=r, column=28).value or '').strip()
                    time_str = str(sheet_main.cell(row=r, column=29).value or '').strip()
                    gold_price_val = sheet_main.cell(row=r, column=13).value

                    try:
                        gold_price = int(extract_digits(str(gold_price_val))) if gold_price_val is not None else 0
                    except ValueError:
                        gold_price = 0

                    datetime_str = f'{date_str} {time_str}'.strip()

                    gold_weight_val = sheet_main.cell(row=r, column=10).value
                    db_gold_weight = None

                    if gold_weight_val is not None:
                        if isinstance(gold_weight_val, (int, float)):
                            db_gold_weight = float(gold_weight_val)
                        else:
                            val_str = str(gold_weight_val).strip().replace('=', '').strip()
                            try:
                                db_gold_weight = float(val_str)
                            except ValueError:
                                pass

                    if db_gold_weight is not None:
                        gold_weight = db_gold_weight
                    else:
                        gold_weight = amount_int / gold_price if gold_price > 0 else 0.0

                    discount = int(round(abs(gold_weight * gold_price - amount_int)))

                    next_export_row = local_idx + 2

                    for col_idx in range(1, 27):
                        val = sheet_main.cell(row=r, column=col_idx).value
                        is_formula = isinstance(val, str) and val.startswith('=')

                        if val is None or is_formula:
                            if col_idx == 1:
                                val = 2
                            elif col_idx == 2:
                                val = start_num + (r - 2)
                            elif col_idx == 3:
                                val = datetime_str
                            elif col_idx == 4:
                                val = 63752299629
                            elif col_idx == 9:
                                val = 1
                            elif col_idx == 10:
                                val = gold_weight
                            elif col_idx == 11:
                                val = 1622
                            elif col_idx == 12:
                                val = discount
                            elif col_idx == 13:
                                val = gold_price
                            elif col_idx == 14:
                                val = 2720000260362
                            elif col_idx in (16, 17, 18, 19):
                                val = 0
                            elif is_formula:
                                val = None

                        target_cell = sheet_export.cell(row=next_export_row, column=col_idx, value=val)

                        if col_idx in (4, 14):
                            target_cell.number_format = '0'

                    sheet_main.cell(row=r, column=31, value='خروجی گرفته شد')
                    exported_count += 1

                if len(chunks) == 1:
                    export_filename = f'tala-export-{time_stamp}.xlsx'
                else:
                    export_filename = f'tala-export-{time_stamp}-part{chunk_idx}.xlsx'

                export_file_path = settings.output_folder / export_filename
                wb_export.save(export_file_path)
                wb_export.close()

            if exported_count > 0:
                try:
                    wb_main.save(self.main_file_path)
                    self._create_database_backup()
                except PermissionError as e:
                    raise RuntimeError('فایل مرکز اصلی طلا باز هست. آن را ببندید و دوباره تلاش کنید..') from e

        finally:
            wb_main.close()

        return exported_count